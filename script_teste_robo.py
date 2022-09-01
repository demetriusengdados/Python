from openpyxl import load_workbook
from openpyxl import Workbook

#Lista de arquivos Excel para juntar
lista_arquivos = ['CustosAutom','PopulacaoPOA','SuperMercado']

#Novo arquivo
wb = Workbook()
nome_arquivo_final = "resultado.xlsx"

for nome_arquivo in lista_arquivos:
    arquivo = load_workbook(filename="%s.xlsx" % nome_arquivo)
    sheet = arquivo[nome_arquivo]
    max_linhas = sheet.max_row
    max_colunas = sheet.max_column
    
    #Nova planilha no resultado.xlsx
    ws = wb.create_sheet(title=nome_arquivo)
    # copying the cell values from source  
    # excel file to destination excel file 
    for i in range (1, max_linhas + 1): 
        for j in range (1, max_colunas + 1): 
            # reading cell value from source excel file 
            c = sheet.cell(row = i, column = j) 
    
            # writing the read value to destination excel file 
            ws.cell(row = i, column = j).value = c.value 

wb.remove(wb['Sheet'])
wb.save(nome_arquivo_final)

    
